from flask import Flask, json, jsonify, redirect, render_template, request, session, make_response, url_for,Response
from flask_cors import CORS, cross_origin
from flask_session import Session
import json
import logging
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.writer.excel import save_virtual_workbook
import pandas as pd
from pytrends.request import TrendReq

# https://stackoverflow.com/questions/42957871/return-a-created-excel-file-with-flask

max_keywords=5

# Set up logging
logging.basicConfig(
  filename='googleqat.log',
  # encoding='utf-8',
  format='%(asctime)s %(levelname)s:%(message)s',
  level=logging.DEBUG
)
logging.debug("Logging activated")

# Create the Flask instance
app = Flask(__name__)

# Enable Cross-Origin Resource Sharing for API use from another IP and/or port
cors = CORS(app)
app.config['CORS_HEADERS'] = 'Content-Type'

# Use Flask server session to avoid a "Confirm Form Resubmission" pop-up:
# Redirect and pass form values from post to get method
app.config['SECRET_KEY'] = "your_secret_key" 
app.config['SESSION_TYPE'] = 'filesystem' 
app.config['SESSION_PERMANENT']= False
app.config.from_object(__name__)
Session(app)

# Extract the Google Trends data for the given keywords
def extract(str):
  # Get the list of keywords
  keywords=str.split(',')
  # Restrict to max_keywords
  keywords=keywords[:max_keywords]  
    
  # Prepare the results
  # Separate dataframes
  df_topics_rising  = pd.DataFrame()
  df_topics_top     = pd.DataFrame()
  df_queries_rising = pd.DataFrame()
  df_queries_top    = pd.DataFrame()
  
  for keyword in keywords:
  
    # payload=[kw]
    # keyword.append(kw)
  
    print(f"extract - keyword: {keyword}")
  
    # Build the pytrends payload
    pytrends.build_payload([keyword])
    
    # Retrieve the related Queries from Google Trends
    # print(f"extract queries - keyword: {keyword}")
    r=pytrends.related_queries()
    d=r[keyword]
    rising=d['rising']
    rising['keyword']=keyword
    df_queries_rising=df_queries_rising.append(rising)
    top=d['top']
    top['keyword']=keyword
    df_queries_top=df_queries_top.append(top)
    
    # Retrieve the related Topics from:was Google Trends
    # print(f"extract topics - keyword: {keyword}")
    r=pytrends.related_topics()
    d=r[keyword]
    rising=d['rising']
    rising['keyword']=keyword
    df_topics_rising=df_topics_rising.append(rising)
    top=d['top']
    top['keyword']=keyword
    df_topics_top=df_topics_top.append(top)
    
    # Move Keyword columns to the first column
    c = df_queries_rising.pop('keyword')
    df_queries_rising.insert(0, 'keyword', c)
    c = df_queries_top.pop('keyword')
    df_queries_top.insert(0, 'keyword', c)
    c = df_topics_rising.pop('keyword')
    df_topics_rising.insert(0, 'keyword', c)
    c = df_topics_top.pop('keyword')
    df_topics_top.insert(0, 'keyword', c)
    
  
  # Put the results into a dictionary

  logging.debug('Topics - rising:')
  logging.debug(df_topics_rising.info())
  logging.debug('Topics - top:')
  logging.debug(df_topics_top.info())
  logging.debug('Queries - rising:')
  logging.debug(df_queries_rising.info())
  logging.debug('Queries - top:')
  logging.debug(df_queries_top.info())
    

  response={
    'topics_rising' : df_topics_rising.to_dict('records'),
    'topics_top' : df_topics_top.to_dict('records'),
    'queries_rising' : df_queries_rising.to_dict('records'),
    'queries_top' : df_queries_top.to_dict('records'),
  }
  
  # Return the results
  return response
# extract

def set_sheet(data,ws):

    c=1
    r=1
      
    # Headers - Queries
    if "Queries" in ws.title:
      ws[get_column_letter(c)+str(r)]='Keyword'
      ws[get_column_letter(c+1)+str(r)]='Value'
      ws[get_column_letter(c+2)+str(r)]='Query'
      
      ws.column_dimensions['A'].width = 40
      ws.column_dimensions['B'].width = 15
      ws.column_dimensions['C'].width = 50

    # Headers - Topics
    else:
      ws[get_column_letter(c)+str(r)]='Keyword'
      ws[get_column_letter(c+1)+str(r)]='Value'
      ws[get_column_letter(c+2)+str(r)]='Title'
      ws[get_column_letter(c+3)+str(r)]='Type'
      
      ws.column_dimensions['A'].width = 40
      ws.column_dimensions['B'].width = 15
      ws.column_dimensions['C'].width = 50
      ws.column_dimensions['D'].width = 30
      
    # Data
    for row in data:
      r+=1
      if "Queries" in ws.title:
        ws[get_column_letter(c)+str(r)]=row['keyword']
        ws[get_column_letter(c+1)+str(r)]=row['value']
        ws[get_column_letter(c+2)+str(r)]=row['query']
      else:
        ws[get_column_letter(c)+str(r)]=row['keyword']
        ws[get_column_letter(c+1)+str(r)]=row['formattedValue']
        ws[get_column_letter(c+2)+str(r)]=row['topic_title']
        ws[get_column_letter(c+3)+str(r)]=row['topic_type']
        
    # Auto-size
    # dims = {}
    # for row in ws.rows:
    #  for cell in row:
    #    if cell.value:
    #      dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    #for col, value in dims.items():
    #  ws.column_dimensions[col].width = value


# set_sheet

# API Endpoint
@app.route('/getgoogleqat', methods=['POST'])
@cross_origin()
def api():
  data = json.loads(request.data)
  # print(data)
  keywords=data['keywords']
  # keywords="France,Germany"
  print(f"GoogleTrends - branch: api - keywords: {keywords}")
  response=extract(keywords)
  print(response)
  return jsonify(response)
# api

# HTML home page
@app.route('/', methods=['GET','POST'])
def slash():

  # The 'extract' button was pressed
  if 'extract' in request.form:
    keywords = request.form["keywords"]
    print(f"GoogleTrends - branch: extract - keywords: {keywords}")
    results=extract(keywords)
  
  # Download Option
  elif 'download' in request.form and 'results' in session:
  
    # Create a workbook
    wb = Workbook()
    
    # Assign the sheets
    wsQueriesRising = wb.active
    wsQueriesRising.title = "Queries-Rising"
    wsQueriesTop = wb.create_sheet("Queries-Top")
    wsTopicsRising = wb.create_sheet("Topics-Rising")
    wsTopicsTop = wb.create_sheet("Topics-Top")
    print(wsQueriesRising.title)
    
    # Get the data
    j=session['results']
    if j:
      results=json.loads(session['results'])
    else:
      results=[]
    
    # Set the sheets
    set_sheet(results['queries_rising'],wsQueriesRising)
    set_sheet(results['queries_top'],wsQueriesTop)
    set_sheet(results['topics_rising'],wsTopicsRising)
    set_sheet(results['topics_top'],wsTopicsTop)
    
    return Response(
      save_virtual_workbook(wb),
      headers={
        'Content-Disposition': 'attachment; filename=sheet.xlsx',
        'Content-type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
      }
    ) 
    
  # Redirect
  if request.method=='POST':
    print("GoogleTrends - branch: redirect")
    # for result in results:
    #   print(f"results 3 - k:{result['keyword']} {result['last']}")
    session['results'] = json.dumps(results)
    return redirect(url_for('slash'))

  # Render
  else:
    print("GoogleTrends - branch: render index.html")
    if 'results' in session:
      j=session['results']
      if j:
        results=json.loads(session['results'])
      else:
        results=[]
    else:
      results=[]
    return render_template("index.html",results=results)
  
# slash

# Set the pytrends API language
pytrends = TrendReq(hl='en-US')

if __name__ == '__main__':
  app.run(host='0.0.0.0', port=5004, debug=True)
